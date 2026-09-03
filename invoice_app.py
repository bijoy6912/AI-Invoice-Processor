import json
import os
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from invoice_engine import process_invoice


# =========================================================
# PROJECT PATHS
# =========================================================

PROJECT_DIR = Path(__file__).resolve().parent

SINGLE_EXCEL = PROJECT_DIR / "invoice_report.xlsx"
BATCH_EXCEL = PROJECT_DIR / "invoice_batch_gui_report.xlsx"

latest_report_path = None


# =========================================================
# VALIDATION
# =========================================================

def validate_invoice(data):
    errors = []
    warnings = []

    line_items = data.get("line_items", [])

    if not data.get("invoice_number"):
        warnings.append("Invoice number is missing.")

    if not data.get("vendor_name"):
        warnings.append("Vendor name is missing.")

    if not data.get("customer_name"):
        warnings.append("Customer name is missing.")

    if not line_items:
        errors.append("No line items found.")

    calculated_subtotal = Decimal("0")

    for index, item in enumerate(line_items, start=1):
        try:
            quantity = Decimal(str(item.get("quantity", 0)))
            unit_price = Decimal(str(item.get("unit_price", 0)))
            amount = Decimal(str(item.get("amount", 0)))

            calculated_amount = quantity * unit_price

            if calculated_amount != amount:
                errors.append(
                    f"Line item {index}: quantity × unit price "
                    f"does not match amount."
                )

            calculated_subtotal += amount

        except (InvalidOperation, TypeError, ValueError):
            errors.append(
                f"Line item {index}: invalid numeric value."
            )

    try:
        subtotal = Decimal(str(data.get("subtotal", 0)))
        tax = Decimal(str(data.get("tax", 0)))
        total = Decimal(str(data.get("total", 0)))

        if calculated_subtotal != subtotal:
            errors.append(
                "Line items total does not match subtotal."
            )

        calculated_total = subtotal + tax

        if calculated_total != total:
            errors.append(
                "Subtotal + tax does not match total."
            )

    except (InvalidOperation, TypeError, ValueError):
        errors.append("Invalid subtotal, tax, or total.")

    if errors:
        status = "REVIEW REQUIRED"
    else:
        status = "VALID"

    return status, errors, warnings


# =========================================================
# EXCEL FORMATTING
# =========================================================

def format_excel_sheet(ws):
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                value_length = len(str(cell.value))
                if value_length > max_length:
                    max_length = value_length
            except Exception:
                pass

        ws.column_dimensions[column_letter].width = min(
            max_length + 2,
            45
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# =========================================================
# SINGLE EXCEL REPORT
# =========================================================

def create_excel_report(data, status, errors, warnings, source_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Report"

    headers = [
        "Source File",
        "Invoice Number",
        "Vendor",
        "Customer",
        "Invoice Date",
        "Due Date",
        "Subtotal",
        "Tax",
        "Total",
        "Status",
        "Errors",
        "Warnings"
    ]

    ws.append(headers)

    ws.append([
        source_file,
        data.get("invoice_number", ""),
        data.get("vendor_name", ""),
        data.get("customer_name", ""),
        data.get("invoice_date", ""),
        data.get("due_date", ""),
        data.get("subtotal", ""),
        data.get("tax", ""),
        data.get("total", ""),
        status,
        "; ".join(errors),
        "; ".join(warnings)
    ])

    format_excel_sheet(ws)

    wb.save(SINGLE_EXCEL)

    return SINGLE_EXCEL


# =========================================================
# BATCH EXCEL REPORT
# =========================================================

def create_batch_excel(results):
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Report"

    headers = [
        "Source File",
        "Invoice Number",
        "Vendor",
        "Customer",
        "Invoice Date",
        "Due Date",
        "Subtotal",
        "Tax",
        "Total",
        "Status",
        "Errors",
        "Warnings"
    ]

    ws.append(headers)

    for result in results:
        data = result["data"]

        ws.append([
            result["source_file"],
            data.get("invoice_number", ""),
            data.get("vendor_name", ""),
            data.get("customer_name", ""),
            data.get("invoice_date", ""),
            data.get("due_date", ""),
            data.get("subtotal", ""),
            data.get("tax", ""),
            data.get("total", ""),
            result["status"],
            "; ".join(result["errors"]),
            "; ".join(result["warnings"])
        ])

    format_excel_sheet(ws)

    wb.save(BATCH_EXCEL)

    return BATCH_EXCEL


# =========================================================
# FILE SELECTION
# =========================================================

def select_invoice():
    file_path = filedialog.askopenfilename(
        title="Select Invoice PDF",
        filetypes=[
            ("PDF Files", "*.pdf"),
            ("All Files", "*.*")
        ]
    )

    if file_path:
        selected_file_var.set(file_path)
        status_var.set("Invoice selected.")


def select_multiple_invoices():
    file_paths = filedialog.askopenfilenames(
        title="Select Invoice PDFs",
        filetypes=[
            ("PDF Files", "*.pdf"),
            ("All Files", "*.*")
        ]
    )

    if file_paths:
        selected_files.clear()
        selected_files.extend(file_paths)

        selected_file_var.set(
            f"{len(file_paths)} PDF file(s) selected."
        )

        status_var.set(
            f"{len(file_paths)} invoice(s) ready."
        )


# =========================================================
# SINGLE INVOICE PROCESSING
# =========================================================

def process_single_invoice():
    global latest_report_path

    pdf_path = selected_file_var.get()

    if not pdf_path or not Path(pdf_path).exists():
        messagebox.showwarning(
            "No Invoice",
            "Please select an invoice PDF first."
        )
        return

    try:
        status_var.set("Processing invoice...")
        root.update_idletasks()

        data = process_invoice(
            pdf_path,
            output_path=PROJECT_DIR / "ocr_invoice_result.json"
        )

        status, errors, warnings = validate_invoice(data)

        report_path = create_excel_report(
            data,
            status,
            errors,
            warnings,
            Path(pdf_path).name
        )

        latest_report_path = report_path

        invoice_number_var.set(
            str(data.get("invoice_number", ""))
        )

        vendor_var.set(
            str(data.get("vendor_name", ""))
        )

        total_var.set(
            str(data.get("total", ""))
        )

        validation_var.set(status)

        if status == "VALID":
            status_var.set(
                "Invoice processed successfully."
            )
        else:
            status_var.set(
                "Invoice processed - review required."
            )

        messagebox.showinfo(
            "Processing Complete",
            f"Invoice processing completed.\n\n"
            f"Status: {status}\n"
            f"Excel report:\n{report_path}"
        )

    except Exception as error:
        status_var.set("Processing failed.")

        messagebox.showerror(
            "Processing Error",
            str(error)
        )


# =========================================================
# BATCH PROCESSING
# =========================================================

def process_batch_invoices():
    global latest_report_path

    if not selected_files:
        messagebox.showwarning(
            "No Invoices",
            "Please select multiple invoice PDFs first."
        )
        return

    results = []

    valid_count = 0
    review_count = 0
    failed_count = 0

    for index, pdf_path in enumerate(
        selected_files,
        start=1
    ):
        status_var.set(
            f"Processing {index}/{len(selected_files)}: "
            f"{Path(pdf_path).name}"
        )

        root.update_idletasks()

        try:
            data = process_invoice(
                pdf_path
            )

            status, errors, warnings = validate_invoice(
                data
            )

            if status == "VALID":
                valid_count += 1
            else:
                review_count += 1

            results.append({
                "source_file": Path(pdf_path).name,
                "data": data,
                "status": status,
                "errors": errors,
                "warnings": warnings
            })

        except Exception as error:
            failed_count += 1

            results.append({
                "source_file": Path(pdf_path).name,
                "data": {},
                "status": "FAILED",
                "errors": [str(error)],
                "warnings": []
            })

    report_path = create_batch_excel(results)

    latest_report_path = report_path

    status_var.set(
        f"Batch complete: {valid_count} valid, "
        f"{review_count} review, "
        f"{failed_count} failed."
    )

    messagebox.showinfo(
        "Batch Processing Complete",
        f"Batch processing completed.\n\n"
        f"VALID: {valid_count}\n"
        f"REVIEW REQUIRED: {review_count}\n"
        f"FAILED: {failed_count}\n\n"
        f"Excel report:\n{report_path}"
    )


# =========================================================
# OPEN EXCEL REPORT
# =========================================================

def open_excel_report():
    if latest_report_path is None:
        if SINGLE_EXCEL.exists():
            report_path = SINGLE_EXCEL
        elif BATCH_EXCEL.exists():
            report_path = BATCH_EXCEL
        else:
            messagebox.showwarning(
                "No Report",
                "No Excel report has been created yet."
            )
            return
    else:
        report_path = latest_report_path

    try:
        os.startfile(str(report_path))
    except Exception as error:
        messagebox.showerror(
            "Open Report Error",
            str(error)
        )


# =========================================================
# GUI
# =========================================================

root = tk.Tk()

root.title(
    "AI Invoice Processor"
)

root.geometry(
    "700x560"
)

root.resizable(
    False,
    False
)


selected_files = []


selected_file_var = tk.StringVar()

invoice_number_var = tk.StringVar(
    value="-"
)

vendor_var = tk.StringVar(
    value="-"
)

total_var = tk.StringVar(
    value="-"
)

validation_var = tk.StringVar(
    value="-"
)

status_var = tk.StringVar(
    value="Ready."
)


title_label = tk.Label(
    root,
    text="AI Invoice Processor",
    font=("Arial", 22, "bold")
)

title_label.pack(
    pady=20
)


subtitle_label = tk.Label(
    root,
    text="OCR → Gemini AI → Validation → Excel",
    font=("Arial", 11)
)

subtitle_label.pack(
    pady=(0, 20)
)


# ---------------------------------------------------------
# File selection
# ---------------------------------------------------------

file_frame = tk.Frame(root)

file_frame.pack(
    padx=30,
    fill="x"
)


browse_button = tk.Button(
    file_frame,
    text="Browse Invoice PDF",
    width=22,
    command=select_invoice
)

browse_button.pack(
    side="left",
    padx=5
)


multiple_button = tk.Button(
    file_frame,
    text="Select Multiple Invoices",
    width=22,
    command=select_multiple_invoices
)

multiple_button.pack(
    side="left",
    padx=5
)


selected_label = tk.Label(
    root,
    textvariable=selected_file_var,
    wraplength=620,
    justify="left"
)

selected_label.pack(
    padx=30,
    pady=15
)


# ---------------------------------------------------------
# Processing buttons
# ---------------------------------------------------------

button_frame = tk.Frame(root)

button_frame.pack(
    pady=10
)


process_button = tk.Button(
    button_frame,
    text="Process Invoice",
    width=22,
    height=2,
    command=process_single_invoice
)

process_button.grid(
    row=0,
    column=0,
    padx=8
)


batch_button = tk.Button(
    button_frame,
    text="Process All Invoices",
    width=22,
    height=2,
    command=process_batch_invoices
)

batch_button.grid(
    row=0,
    column=1,
    padx=8
)


report_button = tk.Button(
    root,
    text="Open Excel Report",
    width=25,
    command=open_excel_report
)

report_button.pack(
    pady=10
)


# ---------------------------------------------------------
# Results
# ---------------------------------------------------------

result_frame = tk.LabelFrame(
    root,
    text="Latest Result",
    padx=20,
    pady=15
)

result_frame.pack(
    padx=30,
    pady=15,
    fill="x"
)


tk.Label(
    result_frame,
    text="Invoice Number:"
).grid(
    row=0,
    column=0,
    sticky="w",
    pady=5
)

tk.Label(
    result_frame,
    textvariable=invoice_number_var
).grid(
    row=0,
    column=1,
    sticky="w",
    padx=20
)


tk.Label(
    result_frame,
    text="Vendor:"
).grid(
    row=1,
    column=0,
    sticky="w",
    pady=5
)

tk.Label(
    result_frame,
    textvariable=vendor_var
).grid(
    row=1,
    column=1,
    sticky="w",
    padx=20
)


tk.Label(
    result_frame,
    text="Total:"
).grid(
    row=2,
    column=0,
    sticky="w",
    pady=5
)

tk.Label(
    result_frame,
    textvariable=total_var
).grid(
    row=2,
    column=1,
    sticky="w",
    padx=20
)


tk.Label(
    result_frame,
    text="Validation:"
).grid(
    row=3,
    column=0,
    sticky="w",
    pady=5
)

tk.Label(
    result_frame,
    textvariable=validation_var
).grid(
    row=3,
    column=1,
    sticky="w",
    padx=20
)


# ---------------------------------------------------------
# Status
# ---------------------------------------------------------

status_label = tk.Label(
    root,
    textvariable=status_var,
    font=("Arial", 10)
)

status_label.pack(
    pady=10
)


root.mainloop()